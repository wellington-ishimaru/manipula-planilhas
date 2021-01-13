import openpyxl
from openpyxl import load_workbook
from datetime import datetime


class Funcionario:
    def __init__(self, nome, numero):
        self.nome = nome
        self.id = numero

    def get_nome(self):
        return self.nome

    def get_id(self):
        return self.id


class Despesas:
    def __init__(self, nome, valor, data):
        self.nome = nome
        self.valor = valor
        self.data = data

    def get_nome(self):
        return self.nome

    def get_valor(self):
        return self.valor

    def get_data(self):
        return self.data


def cadastra_funcionario(nome, numero):
    funcionario = Funcionario(nome, numero)
    return funcionario


def consulta_funcionarios():
    le_planilha("Funcionarios")


def cadastra_despesas(nome, valor, data):
    despesa = Despesas(nome, valor, data)
    return despesa


def consulta_despesas():
    le_planilha("Despesas")


def imprime_menu_opcoes():
    print("******************************************")
    print("********* Bem vindo ao sistema ***********")
    print("******************************************\n")
    print("(1) Cadastrar funcionarios")
    print("(2) Cadastrar despesas")
    print("(3) Consultar funcionarios")
    print("(4) Consultar despesas")
    print("(0) Sair")


def pega_escolha(escolha):
    if escolha == 1:
        nome = input("Informe o nome do funcionário: ")
        lista_funcionarios.append(Funcionario(nome, conta_id))
    elif escolha == 2:
        nome = input("Informe o nome da despesa: ")
        valor = float(input("Informe o valor da despesa: "))
        data = input("Digite a data no formato DD-MM-AAAA: ")
        data = datetime.strptime(data, '%d-%m-%Y')
        lista_despesas.append(Despesas(nome, valor, data))
    elif escolha == 3:
        consulta_funcionarios()
    elif escolha == 4:
        consulta_despesas()
    else:
        print("Escolha inválida")


def cria_pasta(lista_funcionarios, lista_despesas):
    pasta_condominio = openpyxl.Workbook()
    pagina_funcionarios = pasta_condominio.create_sheet(index=0, title='Funcionarios')
    pagina_despesas = pasta_condominio.create_sheet(index=1, title='Despesas')
    pagina_funcionarios["A1"] = "ID"
    pagina_funcionarios["B1"] = "Nome"
    pagina_despesas["A1"] = "Nome"
    pagina_despesas["B1"] = "Valor(R$)"
    pagina_despesas["C1"] = "Data"
    for i in range(len(lista_funcionarios)):
        pagina_funcionarios[f"A{i + 2}"] = lista_funcionarios[i].get_id()
        pagina_funcionarios[f"B{i + 2}"] = lista_funcionarios[i].get_nome()
    for i in range(len(lista_despesas)):
        pagina_despesas[f"A{i + 2}"] = lista_despesas[i].get_nome()
        pagina_despesas[f"B{i + 2}"] = lista_despesas[i].get_valor()
        pagina_despesas[f"C{i + 2}"] = lista_despesas[i].get_data().date()
    pasta_condominio.save("Condominio.xlsx")


def le_planilha(nome):
    planilha = load_workbook(filename="Condominio.xlsx")
    pagina = planilha[nome]
    conta_linha = pagina.max_row
    if nome == "Funcionarios":
        for i in range(1, conta_linha + 1):
            celula_id = str(pagina.cell(row=i, column=1).value)
            celula_nome = pagina.cell(row=i, column=2).value
            print("{:<3} | {:<15}".format(celula_id, celula_nome))
    if nome == "Despesas":
        for i in range(1, conta_linha + 1):
            celula_nome = pagina.cell(row=i, column=1).value
            celula_valor = str(pagina.cell(row=i, column=2).value)
            celula_data = str(pagina.cell(row=i, column=3).value)
            print("{:<15} | {:<10} | {:<10}".format(celula_nome, celula_valor, celula_data))


# main
lista_funcionarios = []
lista_despesas = []
conta_id = 0
imprime_menu_opcoes()
escolha = int(input("Digite a opção desejada: "))
while escolha != 0:
    if escolha == 1:
        conta_id += 1
    pega_escolha(escolha)
    escolha = int(input("Digite a opção desejada: "))
    cria_pasta(lista_funcionarios, lista_despesas)





